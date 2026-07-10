"""
files/tests.py

Test qatlamlari:
  - ProcessingJobModelTests        -> model darajasi
  - ProcessExcelCleanTests         -> services.process_excel_clean
  - ProcessExcelToPdfTests         -> services.process_excel_to_pdf
  - ProcessPdfTableExtractTests    -> services.process_pdf_table_extract
  - ProcessJobTaskTests            -> tasks.process_job (Celery, eager mode)
  - ProcessingJobAPITests          -> views.ProcessingJobViewSet (DRF API)

Eslatma: fayllar BytesIO orqali xotirada yaratiladi (diskka vaqtinchalik fayl
yozilmaydi), shuning uchun Windows'dagi "fayl ikkita joyda ochiq" muammosi
umuman yuzaga kelmaydi va tearDown'da hech narsa tozalash shart emas.
"""

import io

import pandas as pd
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table

from .models import ProcessingJob
from .services import (
    process_excel_clean,
    process_excel_to_pdf,
    process_pdf_table_extract,
)
from .tasks import process_job


def _make_excel_upload(rows, columns, filename="input.xlsx"):
    """Berilgan qatorlardan xotirada .xlsx fayl yaratib, Django upload
    fayl obyektiga aylantiradi. Diskka yozmaydi -> Windows'da xavfsiz."""
    df = pd.DataFrame(rows, columns=columns)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return SimpleUploadedFile(
        filename,
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _make_pdf_upload_with_table(rows, filename="input.pdf"):
    """reportlab yordamida jadval bor oddiy PDF yaratadi (xotirada).
    GRID chizig'i shart -- pdfplumber jadvalni chegara chiziqlari orqali
    aniqlaydi, chiziqsiz Table oddiy matn sifatida ko'rinadi va extract
    funksiyasi uni "jadval yo'q" deb hisoblaydi."""
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    table = Table(rows)
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    doc.build([table])
    buffer.seek(0)
    return SimpleUploadedFile(filename, buffer.read(), content_type="application/pdf")


def _make_empty_pdf_upload(filename="empty.pdf"):
    """Jadvalsiz, faqat matn bor PDF -> extract natijasi bo'sh bo'lishi kerak."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    doc.build([Paragraph("Bu yerda jadval yo'q, faqat matn.", styles["Normal"])])
    buffer.seek(0)
    return SimpleUploadedFile(filename, buffer.read(), content_type="application/pdf")


@override_settings(MEDIA_ROOT=None)  # har testda tmp_path bilan almashtiriladi
class _MediaIsolatedTestCase(TestCase):
    """MEDIA_ROOT'ni har test uchun alohida vaqtinchalik papkaga yo'naltiradi,
    shunda testlar bir-birining natija fayllariga tegmaydi va real media/
    papkasini axlatlamaydi."""

    def setUp(self):
        import tempfile
        self._tmp_dir = tempfile.TemporaryDirectory()
        self._override = override_settings(MEDIA_ROOT=self._tmp_dir.name)
        self._override.enable()
        self.media_root = self._tmp_dir.name

    def tearDown(self):
        self._override.disable()
        self._tmp_dir.cleanup()


class ProcessingJobModelTests(TestCase):
    def test_default_status_is_pending(self):
        job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_CLEAN)
        self.assertEqual(job.status, ProcessingJob.Status.PENDING)

    def test_str_includes_job_type(self):
        job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_TO_PDF)
        self.assertIn("excel_to_pdf", str(job))

    def test_str_includes_status(self):
        job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_TO_PDF)
        self.assertIn("pending", str(job))

    def test_options_defaults_to_empty_dict(self):
        job = ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_CLEAN)
        self.assertEqual(job.options, {})


class ProcessExcelCleanTests(_MediaIsolatedTestCase):
    def test_removes_fully_empty_rows(self):
        upload = _make_excel_upload(
            rows=[["A", 1], [None, None], ["B", 2]],
            columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        result_relpath = process_excel_clean(job)
        result_df = pd.read_excel(f"{self.media_root}/{result_relpath}")

        self.assertEqual(len(result_df), 2)
        self.assertListEqual(list(result_df["name"]), ["A", "B"])

    def test_removes_duplicate_rows_across_all_columns(self):
        upload = _make_excel_upload(
            rows=[["A", 1], ["A", 1], ["B", 2]],
            columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        result_relpath = process_excel_clean(job)
        result_df = pd.read_excel(f"{self.media_root}/{result_relpath}")

        self.assertEqual(len(result_df), 2)

    def test_dedup_by_specific_column_only(self):
        upload = _make_excel_upload(
            rows=[["A", 1], ["A", 2], ["B", 3]],
            columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN,
            input_file=upload,
            options={"dedup_columns": ["name"]},
        )

        result_relpath = process_excel_clean(job)
        result_df = pd.read_excel(f"{self.media_root}/{result_relpath}")

        self.assertEqual(len(result_df), 2)
        self.assertListEqual(sorted(result_df["name"]), ["A", "B"])

    def test_empty_excel_produces_empty_result_without_crashing(self):
        """0 qatorli Excel kelsa, funksiya xato bermasdan bo'sh natija qaytarishi kerak."""
        upload = _make_excel_upload(rows=[], columns=["name", "value"])
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        result_relpath = process_excel_clean(job)
        result_df = pd.read_excel(f"{self.media_root}/{result_relpath}")

        self.assertEqual(len(result_df), 0)

    def test_dedup_column_not_present_raises_key_error(self):
        """options ichida mavjud bo'lmagan ustun nomi berilsa, pandas KeyError
        ko'taradi. Bu hozirgi xatti-harakat -- kelajakda foydalanuvchiga
        tushunarli xabar bilan almashtirish kerak bo'lishi mumkin, lekin
        hozircha shu xatti-harakatni aniq belgilab qo'yamiz."""
        upload = _make_excel_upload(
            rows=[["A", 1], ["B", 2]],
            columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN,
            input_file=upload,
            options={"dedup_columns": ["mavjud_emas"]},
        )

        with self.assertRaises(KeyError):
            process_excel_clean(job)


class ProcessExcelToPdfTests(_MediaIsolatedTestCase):
    def test_creates_pdf_file(self):
        upload = _make_excel_upload(
            rows=[["A", 1], ["B", 2]],
            columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_TO_PDF, input_file=upload,
        )

        result_relpath = process_excel_to_pdf(job)

        self.assertTrue(result_relpath.endswith(".pdf"))
        full_path = f"{self.media_root}/{result_relpath}"
        with open(full_path, "rb") as f:
            header = f.read(5)
        self.assertEqual(header, b"%PDF-")

    def test_empty_excel_still_produces_valid_pdf(self):
        """Ustunlar bor lekin qator yo'q -- PDF baribir yaratilishi kerak
        (jadval faqat header qator bilan)."""
        upload = _make_excel_upload(rows=[], columns=["name", "value"])
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_TO_PDF, input_file=upload,
        )

        result_relpath = process_excel_to_pdf(job)
        full_path = f"{self.media_root}/{result_relpath}"
        with open(full_path, "rb") as f:
            header = f.read(5)
        self.assertEqual(header, b"%PDF-")


class ProcessPdfTableExtractTests(_MediaIsolatedTestCase):
    def test_extracts_table_into_excel_sheet(self):
        upload = _make_pdf_upload_with_table(
            rows=[["name", "value"], ["A", "1"], ["B", "2"]],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.PDF_TABLE_EXTRACT, input_file=upload,
        )

        result_relpath = process_pdf_table_extract(job)
        full_path = f"{self.media_root}/{result_relpath}"

        from openpyxl import load_workbook
        wb = load_workbook(full_path)
        self.assertGreaterEqual(len(wb.sheetnames), 1)
        self.assertNotIn("empty", wb.sheetnames)

    def test_pdf_without_table_produces_empty_sheet(self):
        upload = _make_empty_pdf_upload()
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.PDF_TABLE_EXTRACT, input_file=upload,
        )

        result_relpath = process_pdf_table_extract(job)
        full_path = f"{self.media_root}/{result_relpath}"

        from openpyxl import load_workbook
        wb = load_workbook(full_path)
        self.assertIn("empty", wb.sheetnames)


@override_settings(CELERY_TASK_ALWAYS_EAGER=True, CELERY_TASK_EAGER_PROPAGATES=True)
class ProcessJobTaskTests(_MediaIsolatedTestCase):
    """CELERY_TASK_ALWAYS_EAGER=True bilan task real broker'siz, joriy
    jarayonda sinxron bajariladi -- Redis ishga tushirish shart emas."""

    def test_successful_job_marks_status_done_and_sets_result_file(self):
        upload = _make_excel_upload(
            rows=[["A", 1], ["B", 2]], columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        process_job(str(job.id))
        job.refresh_from_db()

        self.assertEqual(job.status, ProcessingJob.Status.DONE)
        self.assertTrue(bool(job.result_file))
        self.assertEqual(job.error_message, "")

    def test_unknown_job_type_marks_status_failed(self):
        upload = _make_excel_upload(
            rows=[["A", 1]], columns=["name", "value"],
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )
        # Modelning JobType cheklovini chetlab, DB darajasida noto'g'ri qiymat qo'yamiz
        ProcessingJob.objects.filter(id=job.id).update(job_type="noma_lum_tur")
        job.refresh_from_db()

        process_job(str(job.id))
        job.refresh_from_db()

        self.assertEqual(job.status, ProcessingJob.Status.FAILED)
        self.assertIn("noma_lum_tur", job.error_message)

    def test_missing_job_id_does_not_raise(self):
        """Mavjud bo'lmagan job_id bilan chaqirilsa, task jim tugashi kerak
        (xato tashlamasligi kerak) -- hozirgi kod shunday yozilgan."""
        import uuid
        try:
            process_job(str(uuid.uuid4()))
        except Exception as exc:  # noqa: BLE001
            self.fail(f"process_job mavjud bo'lmagan id bilan xato tashladi: {exc}")

    def test_corrupted_input_file_marks_status_failed_with_error_message(self):
        """Fayl kengaytmasi .xlsx lekin ichi haqiqiy Excel emas -- pandas
        o'qishda xato beradi, task buni ushlab FAILED holatiga o'tkazishi kerak."""
        bad_upload = SimpleUploadedFile(
            "broken.xlsx", b"bu Excel fayl emas, oddiy matn", content_type="application/octet-stream",
        )
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=bad_upload,
        )

        process_job(str(job.id))
        job.refresh_from_db()

        self.assertEqual(job.status, ProcessingJob.Status.FAILED)
        self.assertNotEqual(job.error_message, "")


@override_settings(CELERY_TASK_ALWAYS_EAGER=True, CELERY_TASK_EAGER_PROPAGATES=True)
class ProcessingJobAPITests(_MediaIsolatedTestCase, APITestCase):
    def test_create_job_returns_201_and_job_detail(self):
        upload = _make_excel_upload(
            rows=[["A", 1], ["B", 2]], columns=["name", "value"],
        )
        response = self.client.post(
            reverse("job-list"),
            data={"job_type": "excel_clean", "input_file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["job_type"], "excel_clean")
        self.assertIn("id", response.data)

    def test_create_job_with_invalid_job_type_returns_400(self):
        upload = _make_excel_upload(rows=[["A", 1]], columns=["name", "value"])
        response = self.client.post(
            reverse("job-list"),
            data={"job_type": "not_a_real_type", "input_file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_job_without_file_returns_400(self):
        response = self.client.post(
            reverse("job-list"),
            data={"job_type": "excel_clean"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_job_over_size_limit_returns_400(self):
        """serializers.py da 20MB limit bor -- shuni chegara qiymatidan
        oshirib sinaymiz."""
        big_upload = SimpleUploadedFile(
            "big.xlsx", b"0" * (21 * 1024 * 1024), content_type="application/octet-stream",
        )
        response = self.client.post(
            reverse("job-list"),
            data={"job_type": "excel_clean", "input_file": big_upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_retrieve_existing_job_returns_200(self):
        upload = _make_excel_upload(rows=[["A", 1]], columns=["name", "value"])
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        response = self.client.get(reverse("job-detail", args=[job.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["id"], str(job.id))

    def test_retrieve_nonexistent_job_returns_404(self):
        import uuid
        response = self.client.get(reverse("job-detail", args=[uuid.uuid4()]))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_list_jobs_returns_all_created_jobs(self):
        upload1 = _make_excel_upload(rows=[["A", 1]], columns=["name", "value"])
        upload2 = _make_excel_upload(rows=[["B", 2]], columns=["name", "value"])
        ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload1)
        ProcessingJob.objects.create(job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload2)

        response = self.client.get(reverse("job-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 2)

    def test_delete_method_not_allowed(self):
        """ViewSet http_method_names = ["get", "post", "head"] -- DELETE
        ruxsat etilmagan bo'lishi kerak."""
        upload = _make_excel_upload(rows=[["A", 1]], columns=["name", "value"])
        job = ProcessingJob.objects.create(
            job_type=ProcessingJob.JobType.EXCEL_CLEAN, input_file=upload,
        )

        response = self.client.delete(reverse("job-detail", args=[job.id]))

        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)